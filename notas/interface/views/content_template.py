from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook


@login_required
def download_content_master_template(request):
    wb = Workbook()

    # ---------- FOODS ----------
    ws = wb.active
    ws.title = "foods"
    ws.append([
        "food_code",
        "name",
        "protein",
        "carbs",
        "fat",
        "created_by_username",
        "is_active",
    ])
    ws.append([
        "CHICKEN_BREAST",
        "Chicken breast",
        31,
        0,
        3.6,
        request.user.username,
        True,
    ])

    # ---------- MEALS ----------
    ws = wb.create_sheet(title="meals")
    ws.append([
        "meal_code",
        "name",
        "created_by_username",
        "is_public",
        "is_forkable",
        "is_copiable",
        "is_draft",
        "forked_from_code",
        "original_author_username",
    ])

    # ---------- MEAL FOODS ----------
    ws = wb.create_sheet(title="meal_foods")
    ws.append([
        "meal_code",
        "food_code",
        "quantity",
        "order",
    ])

    # ---------- DAILYPLANS ----------
    ws = wb.create_sheet(title="dailyplans")
    ws.append([
        "dailyplan_code",
        "name",
        "created_by_username",
        "is_public",
        "is_forkable",
        "is_copiable",
        "is_draft",
    ])

    # ---------- DAILYPLAN MEALS ----------
    ws = wb.create_sheet(title="dailyplan_meals")
    ws.append([
        "dailyplan_code",
        "meal_code",
        "note",
        "hour",
        "order",
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="notas_content_master.xlsx"'
    )

    wb.save(response)
    return response
